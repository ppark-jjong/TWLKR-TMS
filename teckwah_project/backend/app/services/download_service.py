# backend/app/services/download_service.py
import io
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.dashboard_model import Dashboard
from app.repositories.download_repository import DownloadRepository
from app.utils.logger import log_info, log_error
from app.utils.datetime_helper import format_datetime


class DownloadService:
    """대시보드 데이터 다운로드 서비스"""

    def __init__(self, download_repository: DownloadRepository):
        self.repository = download_repository

    def get_dashboard_data_for_download(
        self, start_time: datetime, end_time: datetime
    ) -> Dict[str, Any]:
        """대시보드 데이터 다운로드용 Excel 파일 생성"""
        try:
            log_info(f"다운로드용 데이터 준비: {start_time} ~ {end_time}")
            
            # 1. 데이터 조회
            dashboards = self.repository.get_dashboard_data_by_create_time(start_time, end_time)
            if not dashboards:
                log_info("다운로드할 데이터가 없습니다")
                return {
                    "success": False, 
                    "message": "다운로드할 데이터가 없습니다",
                    "file_data": None,
                    "file_name": None
                }
            
            # 2. Excel 파일 생성
            file_content, file_name = self._generate_dashboard_xlsx(dashboards, start_time, end_time)
            
            return {
                "success": True,
                "message": f"데이터 준비 완료: {len(dashboards)}건",
                "file_data": file_content,
                "file_name": file_name
            }
            
        except Exception as e:
            log_error(e, "다운로드용 데이터 준비 실패")
            return {
                "success": False,
                "message": f"데이터 준비 중 오류 발생: {str(e)}",
                "file_data": None,
                "file_name": None
            }

    def _generate_dashboard_xlsx(
        self, dashboards: List[Dashboard], start_time: datetime, end_time: datetime
    ) -> Tuple[bytes, str]:
        """대시보드 데이터를 엑셀 파일로 변환"""
        try:
            log_info(f"엑셀 파일 생성 시작: {len(dashboards)}건")
            
            # 1. 워크북 및 워크시트 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "대시보드 데이터"
            
            # 2. 헤더 정의
            headers = [
                "번호", "주문번호", "유형", "상태", "부서", "창고", "SLA", 
                "예상도착시간", "생성시간", "출발시간", "완료시간", 
                "우편번호", "지역", "거리(km)", "소요시간(분)", 
                "주소", "수령인", "연락처", "배송담당자", "담당자연락처", "메모"
            ]
            
            # 3. 헤더 스타일 설정
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style="thin"), 
                right=Side(style="thin"), 
                top=Side(style="thin"), 
                bottom=Side(style="thin")
            )
            
            # 4. 헤더 추가
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # 5. 데이터 추가
            for row_idx, dashboard in enumerate(dashboards, 2):
                # 메모 텍스트 추출
                remarks_text = ""
                if hasattr(dashboard, "remarks") and dashboard.remarks:
                    remarks = sorted(dashboard.remarks, key=lambda r: r.created_at, reverse=True)
                    remarks_text = " | ".join([r.content for r in remarks if r.content])
                
                # 유형 매핑
                type_map = {"DELIVERY": "배송", "RETURN": "회수"}
                
                # 상태 매핑
                status_map = {
                    "WAITING": "대기", 
                    "IN_PROGRESS": "진행중", 
                    "COMPLETE": "완료", 
                    "ISSUE": "이슈",
                    "CANCEL": "취소"
                }
                
                # 데이터 행 추가
                row_data = [
                    row_idx - 1,  # 번호
                    dashboard.order_no,  # 주문번호
                    type_map.get(dashboard.type, dashboard.type),  # 유형
                    status_map.get(dashboard.status, dashboard.status),  # 상태
                    dashboard.department,  # 부서
                    dashboard.warehouse,  # 창고
                    dashboard.sla,  # SLA
                    format_datetime(dashboard.eta) if dashboard.eta else "",  # 예상도착시간
                    format_datetime(dashboard.create_time) if dashboard.create_time else "",  # 생성시간
                    format_datetime(dashboard.depart_time) if dashboard.depart_time else "",  # 출발시간
                    format_datetime(dashboard.complete_time) if dashboard.complete_time else "",  # 완료시간
                    dashboard.postal_code,  # 우편번호
                    dashboard.region,  # 지역
                    dashboard.distance,  # 거리(km)
                    dashboard.duration_time,  # 소요시간(분)
                    dashboard.address,  # 주소
                    dashboard.customer,  # 수령인
                    dashboard.contact,  # 연락처
                    dashboard.driver_name,  # 배송담당자
                    dashboard.driver_contact,  # 담당자연락처
                    remarks_text  # 메모
                ]
                
                # 데이터 추가 및 스타일 설정
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    
                    # 번호, 상태 등은 가운데 정렬
                    if col_idx in [1, 3, 4, 5, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 6. 열 너비 자동 조정
            for col_idx, header in enumerate(headers, 1):
                column_width = max(len(str(header)) * 2, 10)
                # 주소, 메모 열은 더 넓게 설정
                if header in ["주소", "메모"]:
                    column_width = 40
                # 이름, 연락처 등은 적당히 설정
                elif header in ["수령인", "연락처", "배송담당자", "담당자연락처"]:
                    column_width = 15
                # 시간 관련 열은 날짜 형식에 맞게 설정
                elif "시간" in header:
                    column_width = 20
                
                ws.column_dimensions[get_column_letter(col_idx)].width = column_width
            
            # 7. 파일 저장 (메모리 스트림)
            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            
            # 8. 파일명 생성
            file_name = f"대시보드_데이터_{start_time.strftime('%Y%m%d')}_{end_time.strftime('%Y%m%d')}.xlsx"
            
            log_info(f"엑셀 파일 생성 완료: {file_name}")
            return file_stream.getvalue(), file_name
            
        except Exception as e:
            log_error(e, "엑셀 파일 생성 실패")
            raise ValueError(f"엑셀 파일 생성 실패: {str(e)}")
            
    def get_download_date_range(self) -> Dict[str, str]:
        """다운로드 가능한 날짜 범위 조회"""
        try:
            oldest_date, latest_date = self.repository.get_create_time_date_range()
            return {
                "oldest_date": oldest_date.strftime("%Y-%m-%d"),
                "latest_date": latest_date.strftime("%Y-%m-%d")
            }
        except Exception as e:
            log_error(e, "다운로드 가능 날짜 범위 조회 실패")
            # 오류 발생 시 기본값 반환
            now = datetime.now()
            return {
                "oldest_date": (now - timedelta(days=30)).strftime("%Y-%m-%d"),
                "latest_date": now.strftime("%Y-%m-%d")
            }